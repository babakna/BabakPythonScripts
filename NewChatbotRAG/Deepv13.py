# pip install chromadb ollama pypdf2 python-docx openpyxl tk tqdm

import os
import subprocess
import threading
import time
import uuid
import warnings
from tkinter import *
from tkinter import filedialog, messagebox, ttk

import chromadb
import ollama
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader
from tqdm import tqdm

# Suppress cryptography warnings - commenting as userMessage is undefined
# warnings.filterwarnings("ignore", category=UserMessage, module="pypdf")

class RAGApplication:
    def __init__(self, root):
        self.root = root
        self.root.title("RAG Desktop Assistant")
        self.root.geometry("800x600")
        
        self.llm_options = []
        self.current_llm = ""
        self.chroma_client = chromadb.PersistentClient(path="chroma_db")
        self.processing = False
        self.ollama_process = None
        self.uploaded_documents = []
        self.existing_ids = set()

        self.create_widgets()
        self.initialize_chroma()
        self.initialize_models()

    def create_widgets(self):
        # Model Selection Frame
        model_frame = LabelFrame(self.root, text="LLM Selection")
        model_frame.pack(fill=X, padx=10, pady=5)
        
        self.llm_combobox = ttk.Combobox(model_frame, state="readonly", width=40)
        self.llm_combobox.pack(side=LEFT, padx=5)
        
        self.test_btn = Button(model_frame, text="Test Ollama", command=self.test_ollama_connection)
        self.test_btn.pack(side=LEFT, padx=5)
        
        self.refresh_btn = Button(model_frame, text="Refresh Models", command=self.threaded_refresh_models)
        self.refresh_btn.pack(side=LEFT, padx=5)

        # Document Management Frame
        doc_frame = LabelFrame(self.root, text="Document Management")
        doc_frame.pack(fill=X, padx=10, pady=5)
        
        self.upload_btn = Button(doc_frame, text="Upload Documents", command=self.threaded_upload_documents)
        self.upload_btn.pack(side=LEFT, padx=5)
        
        self.show_docs_btn = Button(doc_frame, text="Show Uploaded Docs", command=self.display_uploaded_documents)
        self.show_docs_btn.pack(side=LEFT, padx=5)
        
        self.reset_btn = Button(doc_frame, text="Reset Session", command=self.threaded_reset_system)
        self.reset_btn.pack(side=RIGHT, padx=5)

        # Query Interface Frame
        query_frame = LabelFrame(self.root, text="Query Interface")
        query_frame.pack(fill=BOTH, expand=True, padx=10, pady=5)
        
        self.query_entry = Text(query_frame, height=4)
        self.query_entry.pack(fill=X, pady=5)
        
        self.run_btn = Button(query_frame, text="Run Query", command=self.threaded_run_query)
        self.run_btn.pack(side=LEFT, pady=5)
        
        self.exit_btn = Button(query_frame, text="Exit", command=self.cleanup_and_exit)
        self.exit_btn.pack(side=RIGHT, pady=5)

        # Output Display Frame
        output_frame = LabelFrame(self.root, text="Output")
        output_frame.pack(fill=BOTH, expand=True, padx=10, pady=5)
        
        self.output_text = Text(output_frame, wrap=WORD)
        scrollbar = Scrollbar(output_frame, command=self.output_text.yview)
        self.output_text.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side=RIGHT, fill=Y)
        self.output_text.pack(fill=BOTH, expand=True)

        # Progress Bar
        self.progress = ttk.Progressbar(self.root, orient=HORIZONTAL, mode='indeterminate')
        self.progress.pack(fill=X, padx=10, pady=5)

    def initialize_models(self):
        try:
            if not self.check_ollama_running():
                self.start_ollama_server()
                time.sleep(2)
            
            model_data = ollama.list()
            if not model_data or 'models' not in model_data:
                raise ValueError("Ollama returned unexpected response format")
            
            models = [model['name'] for model in model_data['models'] if 'name' in model]
            
            if not models:
                raise ValueError("No models found in Ollama")
            
            self.llm_options = sorted(list(set(models)), key=lambda x: x.lower())
            self.llm_combobox['values'] = self.llm_options
            
            if self.llm_options:
                self.current_llm = self.llm_options[0]
                self.llm_combobox.set(self.current_llm)
                self.update_output(f"Loaded {len(self.llm_options)} models")
            else:
                raise ValueError("No valid models found")

        except Exception as e:
            self.update_output(f"Model loading error: {str(e)}")
            self.update_output("Please ensure Ollama is running and models are installed")
            self.llm_options = []
            self.llm_combobox['values'] = []
            self.llm_combobox.set('')

    def initialize_chroma(self):
        try:
            self.chroma_client.delete_collection(name="docs")
            self.update_output("Cleared previous database entries")
        except Exception as e:
            self.update_output("No existing database found")
        
        try:
            self.collection = self.chroma_client.create_collection(name="docs")
            self.update_output("Initialized new database")
            self.uploaded_documents.clear()
            self.existing_ids.clear()
        except Exception as e:
            self.update_output(f"Database initialization failed: {str(e)}")
            raise

    def threaded_upload_documents(self):
        if not self.processing:
            threading.Thread(target=self.upload_documents).start()

    def upload_documents(self):
        self.set_processing(True)
        self.update_output("Starting document ingestion...")
        file_paths = filedialog.askopenfilenames(
            filetypes=[("Documents", "*.pdf *.docx *.xlsx")]
        )
        
        if file_paths:
            documents = []
            metadatas = []
            ids = []
            
            try:
                self.update_output(f"Processing {len(file_paths)} new files...")
                self.progress.start()
                
                for file_path in tqdm(file_paths, desc="Processing files"):
                    if file_path in self.uploaded_documents:
                        self.update_output(f"Skipping duplicate: {os.path.basename(file_path)}")
                        continue
                        
                    content = self.read_file(file_path)
                    if content:
                        doc_id = str(uuid.uuid4())
                        documents.append(content)
                        metadatas.append({"source": file_path})
                        ids.append(doc_id)
                        self.uploaded_documents.append(file_path)
                        self.existing_ids.add(doc_id)
                
                if documents:
                    self.collection.add(
                        documents=documents,
                        metadatas=metadatas,
                        ids=ids
                    )
                    self.update_output(f"Added {len(documents)} new documents")
                    self.update_output(f"Total documents in session: {len(self.uploaded_documents)}")
            except Exception as e:
                self.update_output(f"Error during ingestion: {str(e)}")
            finally:
                self.progress.stop()
                self.set_processing(False)

    def read_file(self, file_path):
        try:
            if file_path.endswith('.pdf'):
                reader = PdfReader(file_path)
                return "\n".join([page.extract_text() for page in reader.pages])
            elif file_path.endswith('.docx'):
                doc = Document(file_path)
                return "\n".join([para.text for para in doc.paragraphs])
            elif file_path.endswith('.xlsx'):
                wb = load_workbook(file_path)
                text_content = []
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        text_content.extend([str(cell) for cell in row if cell is not None])
                return "\n".join(text_content)
        except Exception as e:
            self.update_output(f"Error reading {os.path.basename(file_path)}: {str(e)}")
            return None

    def display_uploaded_documents(self):
        self.update_output("\nUploaded Documents:")
        if self.uploaded_documents:
            for idx, doc in enumerate(self.uploaded_documents, 1):
                self.update_output(f"{idx}. {os.path.basename(doc)}")
        else:
            self.update_output("No documents uploaded yet")

    def clear_output(self):
        self.output_text.config(state=NORMAL)
        self.output_text.delete(1.0, END)
        self.output_text.config(state=DISABLED)

    def threaded_run_query(self):
        if not self.processing:
            threading.Thread(target=self.run_query).start()

    def run_query(self):
        if not self.collection or not self.uploaded_documents:
            self.update_output("No documents loaded! Please upload documents first.")
            return
            
        query = self.query_entry.get("1.0", END).strip()
        self.current_llm = self.llm_combobox.get()
        
        if not query:
            self.update_output("Please enter a query")
            return
        
        self.set_processing(True)
        self.update_output("Processing your query... (This may take a moment)")
        
        try:
            results = self.collection.query(
                query_texts=[query],
                n_results=min(3, len(self.uploaded_documents)))
            
            context = "\n\n".join(results['documents'][0])
            
            response = ollama.generate(
                model=self.current_llm,
                prompt=f"Context: {context}\n\nQuestion: {query}\nAnswer:",
                system="You are a helpful assistant. Base your answers strictly on the provided context.",
                stream=False
            )
            
            self.update_output(f"Response ({self.current_llm}):\n{response['response']}")
        except Exception as e:
            self.update_output(f"Error processing query: {str(e)}")
        finally:
            self.set_processing(False)

    def update_output(self, message):
        def safe_update():
            self.output_text.config(state=NORMAL)
            self.output_text.insert(END, message + "\n")
            self.output_text.see(END)
            self.output_text.config(state=DISABLED)
        self.root.after(0, safe_update)

    def set_processing(self, state):
        def safe_state_update():
            self.processing = state
            state_btn = DISABLED if state else NORMAL
            self.upload_btn.config(state=state_btn)
            self.run_btn.config(state=state_btn)
            self.show_docs_btn.config(state=state_btn)
            self.reset_btn.config(state=state_btn)
            
            if state:
                self.progress.start()
            else:
                self.progress.stop()
        self.root.after(0, safe_state_update)

    def cleanup_and_exit(self):
        if self.ollama_process:
            self.ollama_process.terminate()
        self.root.destroy()

    def check_ollama_running(self):
        try:
            ollama.list()
            return True
        except Exception:
            return False

    def start_ollama_server(self):
        try:
            self.ollama_process = subprocess.Popen(["ollama", "serve"], 
                                                  creationflags=subprocess.CREATE_NEW_CONSOLE)
            time.sleep(5)
            return True
        except Exception as e:
            self.update_output(f"Error starting Ollama: {str(e)}")
            return False

    def test_ollama_connection(self):
        self.set_processing(True)
        try:
            if self.check_ollama_running():
                self.update_output("Ollama is running properly!")
                self.initialize_models()
            else:
                self.update_output("Starting Ollama server...")
                if self.start_ollama_server():
                    self.update_output("Ollama started successfully!")
                    self.initialize_models()
                else:
                    self.update_output("Failed to start Ollama!")
        except Exception as e:
            self.update_output(f"Connection error: {str(e)}")
        self.set_processing(False)

    def threaded_refresh_models(self):
        threading.Thread(target=self.initialize_models).start()

    def threaded_reset_system(self):
        if not self.processing:
            threading.Thread(target=self.reset_system).start()

    def reset_system(self):
        self.set_processing(True)
        self.update_output("Initializing new session...")
        self.initialize_chroma()
        self.initialize_models()
        self.update_output("System reset completed")
        self.set_processing(False)

if __name__ == "__main__":
    root = Tk()
    app = RAGApplication(root)
    root.mainloop()